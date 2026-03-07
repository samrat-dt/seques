import io
from datetime import datetime
from typing import Dict, List

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, HRFlowable

from models import Answer, Question


def export_excel(questions: List[Question], answers: Dict[str, Answer]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Security Questionnaire"

    headers = ["#", "Question", "Answer", "Coverage", "Certainty %", "Sources", "Needs Review", "Status"]
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 22

    coverage_colors = {
        "covered": "D1FAE5",
        "partial": "FEF3C7",
        "none": "FEE2E2",
    }

    for row_num, question in enumerate(questions, 2):
        answer = answers.get(question.id)
        if not answer:
            continue

        ws.cell(row=row_num, column=1, value=question.id.upper())
        ws.cell(row=row_num, column=2, value=question.text)
        ws.cell(row=row_num, column=3, value=answer.draft_answer)

        cov_val = answer.evidence_coverage.value
        cov_cell = ws.cell(row=row_num, column=4, value=cov_val.capitalize())
        color = coverage_colors.get(cov_val, "FFFFFF")
        cov_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cov_cell.alignment = Alignment(horizontal="center")

        ws.cell(row=row_num, column=5, value=answer.ai_certainty)
        ws.cell(row=row_num, column=5).alignment = Alignment(horizontal="center")
        ws.cell(row=row_num, column=6, value="; ".join(answer.evidence_sources))
        ws.cell(row=row_num, column=7, value="Yes" if answer.needs_review else "No")
        ws.cell(row=row_num, column=7).alignment = Alignment(horizontal="center")
        ws.cell(row=row_num, column=8, value=answer.status.value.capitalize())

        for col in range(1, 9):
            ws.cell(row=row_num, column=col).alignment = Alignment(
                wrap_text=True, vertical="top"
            )

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 62
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 36
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 12

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def export_pdf(
    questions: List[Question],
    answers: Dict[str, Answer],
    company_name: str = "Vendor",
) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        leftMargin=inch,
        rightMargin=inch,
        topMargin=inch,
        bottomMargin=inch,
    )

    styles = getSampleStyleSheet()
    story = []

    # Title block
    story.append(Paragraph("Security Questionnaire Response", styles["Title"]))
    story.append(
        Paragraph(
            f"{company_name} &bull; {datetime.now().strftime('%B %d, %Y')}",
            styles["Normal"],
        )
    )
    story.append(Spacer(1, 0.3 * inch))

    coverage_labels = {
        "covered": "Covered",
        "partial": "Partial",
        "none": "No Evidence",
    }

    for question in questions:
        answer = answers.get(question.id)
        if not answer:
            continue

        story.append(HRFlowable(width="100%", thickness=0.5, color="#e2e8f0"))
        story.append(Spacer(1, 0.1 * inch))

        story.append(
            Paragraph(
                f"<b>{question.id.upper()}. {question.text}</b>",
                styles["Heading3"],
            )
        )
        story.append(Paragraph(answer.draft_answer, styles["Normal"]))
        story.append(Spacer(1, 0.05 * inch))

        cov_label = coverage_labels.get(answer.evidence_coverage.value, "")
        meta_parts = [
            f"Coverage: {cov_label}",
            f"Certainty: {answer.ai_certainty}%",
            f"Status: {answer.status.value.capitalize()}",
        ]
        if answer.evidence_sources:
            meta_parts.append(f"Sources: {', '.join(answer.evidence_sources)}")

        story.append(Paragraph(f"<i>{' | '.join(meta_parts)}</i>", styles["Normal"]))

        if answer.needs_review:
            story.append(
                Paragraph("<b>[NEEDS REVIEW]</b>", styles["Normal"])
            )

        if answer.suggested_addition:
            story.append(
                Paragraph(
                    f"<i>Suggested addition: {answer.suggested_addition}</i>",
                    styles["Normal"],
                )
            )

        story.append(Spacer(1, 0.15 * inch))

    doc.build(story)
    return buffer.getvalue()
